import json
from openpyxl.styles import Font
import requests
from config_digital import *

import datetime
import time

def getLastDayOfWeek(p_year,p_week):
    firstdayofweek = datetime.datetime.strptime(f'{p_year}-W{int(p_week )- 1}-1', "%Y-W%W-%w").date()
    lastdayofweek = firstdayofweek + datetime.timedelta(days=6.9)
    return lastdayofweek

def getNextSunday(day):
    lvIsoDate=datetime.datetime.fromisoformat(day)
    lvNextSunday=lvIsoDate
    while lvNextSunday.strftime('%a') !='Sun':
       # print ('Current Day for ' + str(lvNextSunday) + ' is ' + lvNextSunday.strftime('%a'))
        lvNextSunday += datetime.timedelta(1)
    return str(lvNextSunday)

def resetSheet(sheetName, workbook,lvMode):
    if sheetName in workbook.sheetnames:
        del workbook[sheetName]
    workbook.create_sheet(sheetName)
    sheet = workbook[sheetName]
    setSheetHeaderRow(sheet,sheetName,lvMode)
    sheet.freeze_panes = 'A2'

def setSheetHeaderRow(sheet, sheetName,lvMode):
    if sheetName == 'Stories':
        if lvMode == 'Sprint':
          sheet.append(['Id','Name','IssueType','Project','Team','Week','Developer','Feature','ModifiedCycleTime','CycleTime','Release','AssociatedUserStory','Effort','Iteration','State','BugsCount','TeamIterationName','IterationStartDate','IterationEndDate','PivotNameDate','DeploymentDate','UATDate','StoryEndDate', 'DistinctReleaseCount'])
        else:
          sheet.append(['Id','Name','Effort','Project','Team','Feature','ModifiedCycleTime','CycleTime','Release','State','BugsCount','IterationReleaseCount','CalendarWeek','StoryEndDate','PivotWeek','DeploymentDate','UATDate'])  
    elif sheetName == 'Releases':
        sheet.append(['Id','Name','EndDate','Team','Week','BuildMaster', 'Status','Effort', 'IsRolledBack'])
    vHeaderFont=Font(size=14,bold=True)
    for cell in sheet["1:1"]:
        cell.font=vHeaderFont

def ifnull(var, returnVar, subField):
  if var is None:
    return returnVar
  return var[subField]

def delete(sheet):
    while(sheet.max_row > 0):
        sheet.delete(1)
    return

def requestHelper(URL):
    try:
        req = requests.get(URL + '&access_token=' + gvTPToken)
    except requests.exceptions.RequestException as e:
        print("Error for URL" + URL)
        return "Error"
    textData=req.text
    try:
        jsonDict=json.loads(textData)
    except:
        print("Error with JSON Decoding for "+URL)
        return "Error"
    return jsonDict
