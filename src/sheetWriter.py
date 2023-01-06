from distutils.command.config import config
from config_digital import *
import openpyxl
from asyncore import loop
from helperFunctions import ifnull, resetSheet, requestHelper, getLastDayOfWeek, getNextSunday
from datetime import *
# vIterations=set()
# vReleases=set()
import csv

def openWorkbook():
   return openpyxl.load_workbook(gvPath)

def saveWorkbook(book):
    book.save(gvPath)

def getTeamIterationDetails(iterationNumber):
    URL="https://drivetime.tpondemand.com/api/v1/TeamIteration/"+str(iterationNumber)+"?format=json&dateformat=iso&include=[Id,Name,StartDate,EndDate,Effort,Team[Name]]"
    return requestHelper(URL)

def getPersonMetrics():
    f= open('VenuAndTeam.csv', 'w', newline='') 
    writer = csv.writer(f)
    dataHeader=['Id','Name','Effort','Project','Team','Feature','State','TeamIteration','EndDate','Developer','Type']
    writer.writerow(dataHeader)
    # Sushma URL="https://drivetime.tpondemand.com/api/v1/Assignments?where=(GeneralUser.Id%20eq%20684)&format=json&dateformat=iso&take=1000"
    URL="https://drivetime.tpondemand.com/api/v1/Assignments?where=(GeneralUser.Id%20in%20(678,686,685))&format=json&dateformat=iso&take=1000"
    json_dict= requestHelper(URL)
    i=1
    for item in json_dict['Items']:
        URL2="https://drivetime.tpondemand.com/api/v1/UserStories?where=(ID%20eq%20"+str(item['Assignable']['Id'])+")&format=json&dateformat=iso&include=[Id,Name,Effort,Project,Iteration[Name],Team[Name],Feature[Name],LeadTime,CycleTime,Release,TeamIteration[Id],EntityState[Name],TeamIteration,EndDate]"
        json_dict2= requestHelper(URL2)
#        print(str(item['Assignable']['Id']) )
        for item2 in json_dict2['Items']:
            if item2['EndDate'] is not None:
                lvDate=item2['EndDate'][0:10]
            else:
                lvDate=""
            data = [str(item['Assignable']['Id']) ,item['Assignable']['Name'] , str(item2['Effort']) ,ifnull(item2['Project'],"null",'Name') , ifnull(item2['Team'],"null",'Name') ,ifnull(item2['Feature'],"null",'Name'),ifnull(item2['EntityState'],"null",'Name'),ifnull(item2['TeamIteration'],"null",'Name'),lvDate,ifnull(item['GeneralUser'],"null",'FullName'),'Story']
            writer.writerow(data)
        URL3="https://drivetime.tpondemand.com/api/v1/Tasks?where=(ID%20eq%20"+str(item['Assignable']['Id'])+")&format=json&dateformat=iso&include=[Id,Name,Team,Effort,Project[Name],TeamIteration[Name],EntityState[Name],EndDate]"
        json_dict3= requestHelper(URL3)
        for item3 in json_dict3['Items']:
            if item3['EndDate'] is not None:
                lvDate=item3['EndDate'][0:10]
            else:
                lvDate=""
            data = [str(item['Assignable']['Id']) , item['Assignable']['Name'] , str(item3['Effort']) , ifnull(item3['Project'],"null",'Name'), ifnull(item3['Team'],"null",'Name') ,"null" , ifnull(item2['EntityState'],"null",'Name'),ifnull(item3['TeamIteration'],"null",'Name'),lvDate,ifnull(item['GeneralUser'],"null",'FullName'),'Task']
            writer.writerow(data)
            

def populateStories(book):
    for loopProjects in gvProjects:
        for loopTeams in gvTeams:
            sheet = book[gvSheetNameStories]
            URL="https://drivetime.tpondemand.com/api/v1/UserStories?where=(Team.Name%20eq%20%27"+loopTeams+"%27)and(Project.ID%20eq%20"+ str(loopProjects) +")and(CreateDate%20gt%20%272022-01-01%27)&include=[Id,Name,Effort,Project,Iteration[Name],Team[Name],Assignments[GeneralUser,Role],Feature[Name],LeadTime,CycleTime,Release,TeamIteration[Id],EntityState[Name],CustomFields]&append=[Bugs-count]&take=1000&orderbydesc=Effort&format=json&dateformat=iso"
            
            while URL != "":
                json_dict= requestHelper(URL)

                for item in json_dict['Items']:
                    vDevEndDate=""
                    vDeploymentDate=""
                    lvModifiedCycleTime=""
                    lvDeveloper=""
                    lvMultipleDevelopers=0
                    lvWeek=""
                    lvLinkId=""
                    lvLinkRelease=""

                    lvLinkId = "=HYPERLINK(\"https://drivetime.tpondemand.com/entity/" + str(item['Id']) + "\", \"" + str(item['Id']) + "\")"

                    if item['Release'] is not None:
                        lvLinkRelease = "=HYPERLINK(\"https://drivetime.tpondemand.com/entity/" + str(item['Release']['Id']) + "\", \"" + str(item['Release']['Id']) + "\")"

                    if item['EndDate'] is not None:                        
                        lvWeek=datetime.fromisoformat(item['EndDate']).strftime("%W")                                                              

                    if item['TeamIteration'] is not None:
                        lvIterationObject=getTeamIterationDetails(item['TeamIteration']['Id'])
                        lvIterationName=lvIterationObject['Name']
                        lvIterationStartDate=lvIterationObject['StartDate'][0:10]
                        lvIterationEndDate=lvIterationObject['EndDate'][0:10]
                        lvPivot = lvIterationName + " : " + lvIterationStartDate + " - " + lvIterationEndDate
                    else:
                        lvIterationName=""
                        lvIterationStartDate=""
                        lvIterationEndDate=""
                        lvPivot =""
                    
                    for assignment in item['Assignments']['Items']:
                        if (assignment['Role']['Name']=="Developer"):
                            if (lvMultipleDevelopers > 0):
                                lvDeveloper=lvDeveloper + " and "
                            lvDeveloper=lvDeveloper + assignment['GeneralUser']['FullName']
                            lvMultipleDevelopers += 1

                    for custom in item['CustomFields']:
                        if (custom['Name']=="DateToUAT" and custom['Value'] is not None):
                            vDevEndDate=custom['Value']
                        elif(custom['Name']=="DateToProd" and custom['Value'] is not None):
                            vDeploymentDate=custom['Value']
                    if (vDevEndDate !="" and vDevEndDate is not None and vDeploymentDate !="" and vDeploymentDate is not None):
                        lvModifiedCycleTime= (datetime.fromisoformat(str(vDeploymentDate)) -  datetime.fromisoformat(str(vDevEndDate))).days
                        if (lvModifiedCycleTime < 0):
                            lvModifiedCycleTime=0

                    row = [lvLinkId, item['Name'],"UserStory",ifnull(item['Project'],"null",'Name'),ifnull(item['Team'],"null",'Name'),lvWeek,lvDeveloper,ifnull(item['Feature'],"null",'Name'), lvModifiedCycleTime,item['CycleTime'], lvLinkRelease,"",item['Effort'],ifnull(item['TeamIteration'],"",'Id'),ifnull(item['EntityState'],"null",'Name'),item['Bugs-Count'],lvIterationName,lvIterationStartDate,lvIterationEndDate, lvPivot ,vDeploymentDate,vDevEndDate,item['EndDate'], "=IFERROR(1/COUNTIF($I:$I,@$I:$I), 0)"]
                    
                    sheet.append(row)
                   
                try:
                    URL=json_dict['Next']
                except:
                    URL=""


def populateBugs(book):
    for loopProjects in gvProjects:
        for loopTeams in gvTeams:
            sheet = book[gvSheetNameStories]
            URL="https://drivetime.tpondemand.com/api/v1/Bugs?where=(Team.Name%20eq%20%27"+loopTeams+"%27)and(Project.ID%20eq%20"+ str(loopProjects) +")and(CreateDate%20gt%20%272022-01-01%27)&include=[Id,Name,Effort,Project,Iteration[Name],Team[Name],Assignments[GeneralUser,Role],Feature[Name],LeadTime,CycleTime,Release,TeamIteration[Id],EntityState[Name],UserStory[Id],CustomFields]&take=1000&orderbydesc=Effort&format=json&dateformat=iso"

            while URL != "":
                json_dict= requestHelper(URL)

                for item in json_dict['Items']:
                    vDevEndDate=""
                    vDeploymentDate=""
                    lvModifiedCycleTime=""
                    lvDeveloper=""
                    lvMultipleDevelopers=0
                    lvWeek=""
                    lvLinkId=""
                    lvLinkRelease=""
                    lvRelatedUserStory=""

                    lvLinkId = "=HYPERLINK(\"https://drivetime.tpondemand.com/entity/" + str(item['Id']) + "\", \"" + str(item['Id']) + "\")"

                    if item['Release'] is not None:
                        lvLinkRelease = "=HYPERLINK(\"https://drivetime.tpondemand.com/entity/" + str(item['Release']['Id']) + "\", \"" + str(item['Release']['Id']) + "\")"

                    if item['UserStory'] is not None:
                        lvRelatedUserStory = "=HYPERLINK(\"https://drivetime.tpondemand.com/entity/" + str(item['UserStory']['Id']) + "\", \"" + str(item['UserStory']['Id']) + "\")"

                    if item['EndDate'] is not None:                        
                        lvWeek=datetime.fromisoformat(item['EndDate']).strftime("%W")  
                
                    if item['TeamIteration'] is not None:
                        lvIterationObject=getTeamIterationDetails(item['TeamIteration']['Id'])
                        lvIterationName=lvIterationObject['Name']
                        lvIterationStartDate=lvIterationObject['StartDate'][0:10]
                        lvIterationEndDate=lvIterationObject['EndDate'][0:10]
                        lvPivot = lvIterationName + " : " + lvIterationStartDate + " - " + lvIterationEndDate
                    else:
                        lvIterationName=""
                        lvIterationStartDate=""
                        lvIterationEndDate=""
                        lvPivot =""

                    for assignment in item['Assignments']['Items']:
                        if (assignment['Role']['Name']=="Developer"):
                            if (lvMultipleDevelopers > 0):
                                lvDeveloper=lvDeveloper + " and "
                            lvDeveloper=lvDeveloper + assignment['GeneralUser']['FullName']
                            lvMultipleDevelopers += 1

                    for custom in item['CustomFields']:
                        if (custom['Name']=="DateToUAT" and custom['Value'] is not None):
                            vDevEndDate=custom['Value']
                        elif(custom['Name']=="DateToProd" and custom['Value'] is not None):
                            vDeploymentDate=custom['Value']
                    if (vDevEndDate !="" and vDevEndDate is not None and vDeploymentDate !="" and vDeploymentDate is not None):
                        lvModifiedCycleTime= (datetime.fromisoformat(str(vDeploymentDate)) -  datetime.fromisoformat(str(vDevEndDate))).days
                        if (lvModifiedCycleTime < 0):
                            lvModifiedCycleTime=""
#                    
                    row = [lvLinkId, item['Name'],"Bug",ifnull(item['Project'],"null",'Name'),ifnull(item['Team'],"null",'Name'),lvWeek, lvDeveloper,ifnull(item['Feature'],"null",'Name'), lvModifiedCycleTime,item['CycleTime'], lvLinkRelease,lvRelatedUserStory,item['Effort'],ifnull(item['TeamIteration'],"",'Id'),ifnull(item['EntityState'],"null",'Name'),1,lvIterationName,lvIterationStartDate,lvIterationEndDate, lvPivot ,vDeploymentDate,vDevEndDate,item['EndDate'],"=IFERROR(1/COUNTIF($I:$I,@$I:$I), 0)"]
                    
                    sheet.append(row)
                    
                try:
                    URL=json_dict['Next']
                except:
                    URL=""


def populateReleases(book):
    for loopTeams in gvTeams:
        sheet = book[gvSheetNameReleases]
        URL="https://drivetime.tpondemand.com/api/v2/Releases?where=(Teams=%27"+loopTeams+"%27)&select={Id,Name,EndDate,Effort,BuildMaster,rolledback,teams,status}&format=json&dateformat=iso&take=1000"
        json_dict= requestHelper(URL)
        if json_dict != "Error":
            for item in json_dict['items']:
                lvWeek=""
                lvLinkRelease=""

                lvLinkRelease = "=HYPERLINK(\"https://drivetime.tpondemand.com/entity/" + str(item['id']) + "\", \"" + str(item['id']) + "\")"

                if item['endDate'] is not None:                        
                    lvWeek=datetime.fromisoformat(item['endDate']).strftime("%W")
                     
                row = [lvLinkRelease,item['name'],item['endDate'][0:10],item['teams'],lvWeek,item['buildMaster'],item['status'],item['effort'],item['rolledback']]
                sheet.append(row)