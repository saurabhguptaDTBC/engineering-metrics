from distutils.command.config import config
from config_core import *
import openpyxl
from asyncore import loop
from helperFunctions import ifnull, resetSheet, requestHelper, getLastDayOfWeek, getNextSunday
from datetime import *

# vIterations=set()
# vReleases=set()

def openWorkbook():
   return openpyxl.load_workbook(gvPath)

def saveWorkbook(book):
    book.save(gvPath)

def getTeamIterationDetails(iterationNumber):
    URL="https://drivetime.tpondemand.com/api/v1/TeamIteration/"+str(iterationNumber)+"?format=json&dateformat=iso&include=[Id,Name,StartDate,EndDate,Effort,Team[Name]]"
    return requestHelper(URL)


def populateStories(book):
    resetSheet(gvSheetNameStories,book,"kb")
    for loopProjects in gvProjects:
        for loopTeams in gvTeams:
            sheet = book[gvSheetNameStories]
            URL="https://drivetime.tpondemand.com/api/v1/UserStories?where=(Team.Name%20eq%20%27"+loopTeams+"%27)and(Project.ID%20eq%20"+ str(loopProjects) +")and(CreateDate%20gt%20%272022-01-01%27)&include=[Id,Name,Effort,Project,Iteration[Name],Team[Name],Feature[Name],LeadTime,CycleTime,Release,TeamIteration[Id],EntityState[Name],CustomFields]&append=[Bugs-count]&take=1000&orderbydesc=Effort&format=json&dateformat=iso"
            while URL != "":
                json_dict= requestHelper(URL)
                iterationReleaseList=[]
                vIterationRelease=""
                vReleaseVar=""
                releaseCount=0
                lvLastdayofweek=""
                for item in json_dict['Items']:
                     vDevEndDate=""
                     vDeploymentDate=""
                     lvModifiedCycleTime=0
                # The following block is added to get unique releases in iteration : excel wasn't giving me an easy way to do this
                     if item['EndDate'] is not None:
                        #lvWeek=datetime.fromisoformat(item['EndDate']).isocalendar()[1]
                        lvLastdayofweek=getNextSunday(item['EndDate'])[0:10]
                        lvWeek=datetime.fromisoformat(item['EndDate']).strftime("%W")
                        lvEndDate=item['EndDate'][0:10]
                        vReleaseVar=lvLastdayofweek
                     else:
                        lvWeek=""
                        lvEndDate=""
                        lvLastdayofweek=""
                        vReleaseVar="null"
                     if ifnull(item['Release'],"null",'Id') != "null" and vReleaseVar != "null":
                        vIterationRelease=str(ifnull(item['Release'],"null",'Id'))+":"+str(vReleaseVar)
                        if vIterationRelease not in iterationReleaseList:
                            iterationReleaseList.append (vIterationRelease)   
                            releaseCount=1
                        else:
                            releaseCount=0
                     else:
                        releaseCount=0
                 #   if item['TeamIteration'] is not None:
                 #       lvIterationObject=getTeamIterationDetails(item['TeamIteration']['Id'])
                 #       lvIterationName=lvIterationObject['Name']
                 #       lvIterationStartDate=lvIterationObject['StartDate'][0:10]
                 #       lvIterationEndDate=lvIterationObject['EndDate'][0:10]
                 #   else:
                 #       lvIterationName=""
                 #       lvIterationStartDate=""
                 #   
                     for custom in item['CustomFields']:
                        if (custom['Name']=="DateToUAT" and custom['Value'] is not None):
                            vDevEndDate=custom['Value']
                        elif(custom['Name']=="DateToProd" and custom['Value'] is not None):
                            vDeploymentDate=custom['Value']
                     if (vDevEndDate !="" and vDevEndDate is not None and vDeploymentDate !="" and vDeploymentDate is not None):
                        lvModifiedCycleTime= (datetime.fromisoformat(str(vDeploymentDate)) -  datetime.fromisoformat(str(vDevEndDate))).days
                        if (lvModifiedCycleTime < 0):
                            lvModifiedCycleTime=0    
                     #row = [item['Id'], item['Name'],item['Effort'],ifnull(item['Project'],"null",'Name'),ifnull(item['Team'],"null",'Name'),ifnull(item['Feature'],"null",'Name'), item['LeadTime'],item['CycleTime'], ifnull(item['Release'],"",'Id'),ifnull(item['EntityState'],"null",'Name'),item['Bugs-Count'],releaseCount,lvWeek,lvEndDate,lvLastdayofweek]
                     row = [item['Id'], item['Name'],item['Effort'],ifnull(item['Project'],"null",'Name'),ifnull(item['Team'],"null",'Name'),ifnull(item['Feature'],"null",'Name'), lvModifiedCycleTime,item['CycleTime'], ifnull(item['Release'],"",'Id'),ifnull(item['EntityState'],"null",'Name'),item['Bugs-Count'],releaseCount,lvWeek,lvEndDate,lvLastdayofweek,vDeploymentDate,vDevEndDate]
                     sheet.append(row)
                    # if item['Release'] is not None:
                        # vReleases.add(item['Release']['Id'])
                    #if item['TeamIteration'] is not None
                        # vIterations.add(item['TeamIteration']['Id'])
                    #    lvIterationObject=getTeamIterationDetails(item['TeamIteration']['Id'])
                try:
                    URL=json_dict['Next']
                except:
                    URL=""



def populateReleases(book):
    resetSheet(gvSheetNameReleases,book)
    sheet=book[gvSheetNameReleases]
    # for i in vReleases:
    #     sheet = book[vSheetNameReleases]
    #     URL="https://drivetime.tpondemand.com/api/v1/Releases/"+str(i)+"?include=[Name,EndDate,Effort,Owner]&format=json&dateformat=iso"
    #     json_dict= requestHelper(URL)
    #     if json_dict != "Error":
    #         row = [json_dict['Id'],json_dict['Name'],json_dict['EndDate'],json_dict['Effort'],json_dict['Owner']['FullName']]
    #         print(i)
    #         sheet.append(row)
    for loopProjects in gvProjects:
        URL="https://drivetime.tpondemand.com/api/v1/Releases?where=(Projects.ID%20eq%20"+str(loopProjects)+")&include=[Name,EndDate,Effort,Owner]&format=json&dateformat=iso&take=1000"
        json_dict= requestHelper(URL)
        if json_dict != "Error":
            for item in json_dict['Items']:
                row = [item['Id'],item['Name'],item['EndDate'],item['Effort'],item['Owner']['FullName']]
                sheet.append(row)